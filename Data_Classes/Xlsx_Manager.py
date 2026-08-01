# =============================================================================================== #
#                 manages all  lists  for   xlsx  rows                                            #
# =============================================================================================== #
# Xlsx_Rows_From_Sheet_normalized   nRow    Contab  Valuta    Des1    Accr   Addeb     Des2       #
# Xlsx_Rows_Desc_CompactnRow        Contab  Valuta  Des1Comp  Accr   Addeb    Des2Comp            #
# tXlsx_Rows_Compact                nRow  _Contab   Valuta   Accr     Addeb  Full_Desc            #
#                                                                                                 #
# self._tWith_Code_Tree_List        nRow Contabile _Valuta Accred _Addeb TRdesc TRcode            #
# self._tWihtout_Code_Tree_List     nRow Contabile _Valuta Accred _Addeb FullDesc                 #
# ----------------------------------------------------------------------------------------------- #

import warnings
from openpyxl import load_workbook
from Data_Classes.Codes_DB import *

class Xlsx_Manager(Codes_db):
    def __init__(self):
        super().__init__()

        #ignore  : warn("Workbook contains no stylesheet, using openpyxl's defaults")
        warnings.simplefilter("ignore")

        self.Dummy   = 0

        # one row on xlsx file   ---------------------------------------------------------------
        #  nRow timeContab timeValuta  Des1  Accred  Addeb  Des2

        # -------   _tAtt : temporary attributes  that will be copied on _Att  if all OK  -------------
        self.SheetName                         = ""
        self._Xlsx_Rows_From_Sheet_normalized  = []
        self._tXlsx_Rows_Compact      = []     # temporaries to not dommage the _list
        self.tXlsx_Rows_for_view      = []
        self._tWith_Code_Tree_List    = []
        self._tWihtout_Code_Tree_List = []

        self._With_Code_Tree_List    = []      # updated from _t after a successfull Load
        self._Wihtout_Code_Tree_List = []
        self._Records_ToInsert_List  = []

        # ------------------------
        self._tTot_Rows        = 0
        self._tTot_OK          = 0
        self._tTot_NOK         = 0
        self._tTotWith_Code    = 0
        self._tTotWithout_Code = 0
        self._tiYear_List      = []


        # ------------------------------------------------------------ #
        #  -----  the values  are filled  from  Transact_DB.py  ------ #
        #   but they MUST be here because they are used inside here    #
        self._tTransact_Year   = None
        self._tTransact_Table  = []
        # ------------------------------------------------------------ #

        self._Transact_Filename              = ""
        self._Db_Transact_Database           = []
        self.Fetched_List                    = []

    # ----------------------------------------------------------------------------------- #
    #            ----------------      public   methods   -----------------               #
    # ----------------------------------------------------------------------------------- #
    # ----------------------------------------------------------------------------------------------- #
    def Load_Xlsx_Rows(self) -> tuple[bool, str | list]:
        self._Init_Xlsx_Data()
        self._Set_Xlsx_Conto_Year_Month()
        if self._tXlsx_Year is None or self._tXlsx_Conto is None or self._tXlsx_Month is None:
            return False, "FATAL ERROR 13\non extracting Year, Conto , Month from xlsx file"

        Filename = self.Get_sel_dictionary_value(XLSX_FILENAME)
        if not Gl_Cek_Xlsx_Name(Filename):
            return False, "FATAL ERROR 12:\nxlsx filename not OK"

        status, data = self._get_work_sheet_rows_normalized()
        if not status:
            return False, data

        # Xlsx_Rows_From_Sheet_normalized :  nRow  Contab  Valuta  Des1  Accr  Addeb  Des2
        for row in self._tXlsx_Rows_From_Sheet_normalized:
            Checked_Row = self._Check_Values(row)
            if len(Checked_Row) != 0:
                Des1      = Checked_Row[IX_SHEET_DESCR1]
                Des2      = Checked_Row[IX_SHEET_DESCR2]
                Des1_Comp = Compact_Descr_String(Des1)
                Des2_Comp = Compact_Descr_String(Des2)
                Full_Desc  = FullDescr_Setup(Des1_Comp, Des2_Comp)
                row_to_view =  [ Checked_Row[IX_SHEET_NROW],
                                Checked_Row[IX_SHEET_CONTAB],
                                Checked_Row[IX_SHEET_VALUTA],
                                Des1,
                                Checked_Row[IX_SHEET_ACCRED],
                                Checked_Row[IX_SHEET_ADDEB],
                                Des2]
                self._tXlsx_Rows_for_view.append(row_to_view)

                Row_Compact = [ Checked_Row[IX_SHEET_NROW],
                                Checked_Row[IX_SHEET_CONTAB],
                                Checked_Row[IX_SHEET_VALUTA],
                                Checked_Row[IX_SHEET_ACCRED],
                                Checked_Row[IX_SHEET_ADDEB],
                                Full_Desc ]
                self._tXlsx_Rows_Compact.append(Row_Compact)
                pass

        if self._tXlsx_Conto == FLASH or self._tXlsx_Conto == AMBRA or self._tXlsx_Conto == POSTA:
            self._tXlsx_Rows_Compact.sort(reverse=True)

        status, data = self._create_With_Out_codes_lists()
        if not status:
            return False, data
        self._Save_Xlsx_Data()
        return True, ''

    # ------------------------------------------------------------------------------------
    def Get_Length_Xlsx(self):
        return len(self._tXlsx_Rows_Compact)

    def Get_Xlsx_rows_for_view(self):
        return self._Xlsx_Rows_for_view

    # --------------------------------------------------------------------------------------------
    def _Set_Xlsx_Conto_Year_Month(self):
        Filename = self.Get_sel_dictionary_value(XLSX_FILENAME)
        self._tXlsx_Conto = None
        self._tXlsx_Year = None
        self._tXlsx_Month = None
        FullFilename = Filename
        if FullFilename != UNKNOWN:
            filename = Get_File_Name(FullFilename)
            self._tXlsx_Conto = filename[0:5]
            self._tXlsx_Year = int(filename[6:10])
            self._tXlsx_Month = int(filename[11:13])
        pass

    # --------------------------------------------------------------------------------------------
    def _Init_Xlsx_Data(self):
        self._tXlsx_Rows_From_Sheet_normalized = []     # exactly as in datasheet
        self._tXlsx_Rows_for_view              = []     # as in xlsx  None purged
        self._tXlsx_Rows_Compact      = []
        self._tWith_Code_Tree_List    = []
        self._tWihtout_Code_Tree_List = []

        # ------------------------
        self._tTot_Rows        = 0
        self._tTot_OK          = 0
        self._tTot_NOK         = 0
        self._tTotWith_Code    = 0
        self._tTotWithout_Code = 0
        self._tiYear_List = []

        # -----------------------------------------------------------------------------------------
        self._tXlsx_Year     = None  # they are  calculated on startup
        self._tXlsx_Month    = None
        self._tXlsx_Conto    = None  # or on selecting new file  FIDEU_2024_01.xlsx

    # --------------------------------------------------------------------------------------------
    def _Save_Xlsx_Data(self):
        self._Xlsx_Rows_From_Sheet_normalized = self._tXlsx_Rows_From_Sheet_normalized
        self._Xlsx_Rows_for_view              = self._tXlsx_Rows_for_view
        self._Xlsx_Rows_Compact      = self._tXlsx_Rows_Compact

        # ------------------------
        self._Tot_Rows        = self._tTot_Rows
        self._Tot_OK          = self._tTot_OK
        self._Tot_NOK         = self._tTot_NOK
        self._TotWith_Code    = self._tTotWith_Code
        self._TotWihtout_Code = self._tTotWithout_Code
        self._iYear_List      = self._tiYear_List

        self._Xlsx_Conto = self._tXlsx_Conto
        self._Xlsx_Year  = self._tXlsx_Year
        self._Xlsx_Month = self._tXlsx_Month

        self._With_Code_Tree_List    = self._tWith_Code_Tree_List
        self._Wihtout_Code_Tree_List = self._tWihtout_Code_Tree_List
        pass

    # --------------------------------------------------------------------------------- #
    #  Workbook is the container of all Worksheets                                      #
    #  while the Worksheet is the container of Data of one Sheet                        #
    # --------------------------------------------------------------------------------- #
    def _Get_Work_Sheet_Rows(self) -> tuple[bool, str]:
        filename = self.Get_sel_dictionary_value(XLSX_FILENAME)
        Work_Book = None
        try:
            Work_Book = load_workbook(filename)
        except Exception as e:
            Err = f"Error on Worksheet loading\n  {e}"
            self._tTot_Rows = -1
            return False, Err
        finally:
            self.SheetName = Work_Book.sheetnames[0]  # always the first sheet
            self._Work_Sheet = Work_Book[self.SheetName]
            self._tTot_Rows = self._Work_Sheet.max_row
            return True, ''

    # -----------------------------------------------------------------------------------
    def _get_work_sheet_rows_normalized(self) -> tuple[bool, str]:
        status, data = self._Get_Work_Sheet_Rows()
        if not status:
            return False, data
        if self._tTot_Rows == 0:
            return False, f"FATAL ERROR 61:\nil file xlsx non contiene nessuna riga!",
        Des1    = None
        Accred  = None
        Addeb   = None
        Des2    = None
        # CONTO     A       B       C       D       E       F       G
        #           0       1       2       3       4       5       6
        # FIDEU     Contab  Valuta  Des1    Accred  Added   Des2
        # FLH-AMBR  Contab  Valuta  Des1            Addeb-          Accred
        # POSTA     Contab  Valuta  Addeb-  Accred  Des1
        # FORM nRow,Contab, Valuta, Des1,   Accred, Addeb,  Des2
        #        0     1       2      3        4      5       6
        for nRow in range(1, self._tTot_Rows+1):
            Contab = self._Work_Sheet['A' + str(nRow)].value
            Valuta = self._Work_Sheet['B' + str(nRow)].value

            if self._tXlsx_Conto == FIDEU:
                Des1   = self._Work_Sheet['C' + str(nRow)].value
                Accred = self._Work_Sheet['D' + str(nRow)].value
                Addeb  = self._Work_Sheet['E' + str(nRow)].value
                Des2   = self._Work_Sheet['F' + str(nRow)].value

            elif self._tXlsx_Conto == FLASH or self._tXlsx_Conto == AMBRA:
                Des1   = self._Work_Sheet['C' + str(nRow)].value
                Accred = self.float_check(self._Work_Sheet['D' + str(nRow)].value, True)
                Addeb  = self.float_check(self._Work_Sheet['G' + str(nRow)].value, True)
            elif self._tXlsx_Conto == POSTA:
                pass
            self._tXlsx_Rows_From_Sheet_normalized.append([nRow, Contab, Valuta, Des1, Accred, Addeb, Des2])
        return True, ''

    # --------------------------------------------------------------------------------------------- #
    # _With_Code_Tree_List   : nRow   Contab  Valuta  TR_Desc   Accred  Addeb   TRcode  RowFullDes 
    def _create_With_Out_codes_lists(self) -> tuple[bool, str]:
        self._tWihtout_Code_Tree_List = []
        self._tWith_Code_Tree_List    = []
        self._tTotWith_Code           = 0
        self._tTotWithout_Code        = 0

        for Row in self._tXlsx_Rows_Compact:
            Full_Desc  = Row[IX_ROW_COMP_FULLDES]
            TRcodeList  = self._Find_StrToFind_InFullDesc(Full_Desc)
            nCode = len (TRcodeList)
            if nCode == 1:                          # Unic code found for Row
                self.Insert_On_WithCode_List(Row, TRcodeList[0])

            elif nCode > 1:                         # Multiple codes found for a xlsx row
                # print(TRcodeList, Row[IX_ROW_COMP_FULLDES])
                message  = f"la descrizione completa per la riga n.  {str(Row[IX_ROW_COMP_NROW])}\n\n"
                message += f"Contab: {Row[IX_ROW_COMP_CONTAB]}\nValuta: {Row[IX_ROW_COMP_VAL]}\n"
                message += f"Accred: {Row[IX_ROW_COMP_ACCR]}\nAddeb: {Row[IX_ROW_COMP_ADDEB]}\n"
                message += f"{Row[IX_ROW_COMP_FULLDES]}\n\n"
                message += f"combacia con piu stringhe per la ricerca:\n\n"
                for code in TRcodeList:
                    strFound =  f"codice: {str(code)}:  descr: {self.Get_TrDesc_FromCode(code)}\n{self.Get_strToFind_FromCode(code)}\n\n"
                    message += strFound        # dlg_select = View_Messa
                return False, message

            else:  # Code NOT found
                myRow = [Row[IX_ROW_COMP_NROW], self._tXlsx_Conto, Row[IX_ROW_COMP_CONTAB], Row[IX_ROW_COMP_VAL],
                         Row[IX_ROW_COMP_ACCR], Row[IX_ROW_COMP_ADDEB], Row[IX_ROW_COMP_FULLDES] ]

                self._tWihtout_Code_Tree_List.append(myRow)
                self._tTotWithout_Code += 1
                pass
        return True, ''

    # -----------------------------Code(TRcode)---------------------------------------------------------------
    def Insert_On_WithCode_List(self, Row, TRcode):
        TRdesc = self.Get_TrDesc_FromCode(TRcode)
        RecForIns   = [Row[IX_ROW_COMP_NROW], self._tXlsx_Conto, Row[IX_ROW_COMP_CONTAB], Row[IX_ROW_COMP_VAL],
                       Row[IX_ROW_COMP_ACCR], Row[IX_ROW_COMP_ADDEB],
                       TRdesc, TRcode,
                       Row[IX_ROW_COMP_FULLDES] ]
        self._tWith_Code_Tree_List.append(RecForIns)
        self._tTotWith_Code += 1
        pass
    pass

    # ---------------------------------------------------------------------------------------------
    def _Check_Values(self, Xlsx_row):
        Xlsx_Row_List_Checked = []
        nRow = Xlsx_row[IX_SHEET_NROW]
        if not type(nRow is int):
            return []
        Xlsx_Row_List_Checked.append(nRow)

        Contab = Xlsx_row[IX_SHEET_CONTAB]
        if isinstance(Contab, datetime):
            Xlsx_Row_List_Checked.append(Contab)
        else:
            return []

        Valuta = Xlsx_row[IX_SHEET_VALUTA]
        if isinstance(Valuta, datetime):
            Xlsx_Row_List_Checked.append(Valuta)
        else:
            return []

        str_date_contab = Contab.strftime("%Y-%m-%d")
        str_date_valuta = Valuta.strftime("%Y-%m-%d")
        iYear_Contab = int(str_date_contab[0:4])
        iYear_Valuta = int(str_date_valuta[0:4])
        if iYear_Contab != self._tXlsx_Year and iYear_Valuta != self._tXlsx_Year:
            return []

        Descr1 = Xlsx_row[IX_SHEET_DESCR1]  # the first description must be "dddd"
        Descr2 = Xlsx_row[IX_SHEET_DESCR2]
        if type(Descr1) is not str:
            Descr1 = ''
        if type(Descr2) is not str:
            Descr2 = ''
        if Descr1 == '' and Descr2 == '':   # at least one description mus te filled string
                return []
        Xlsx_Row_List_Checked.append(Descr1)

        # this values will be always float
        Accred = self._Convert_Str_To_Float(Xlsx_row[IX_SHEET_ACCRED])
        Xlsx_Row_List_Checked.append(Accred)
        Addeb  = self._Convert_Str_To_Float(Xlsx_row[IX_SHEET_ADDEB])
        Xlsx_Row_List_Checked.append(Addeb)

        Xlsx_Row_List_Checked.append(Descr2)
        return Xlsx_Row_List_Checked

    # ---------------------------------------------------------------------------------------
    def float_check(self, value, negate):
        self.Dummy = 0
        if isinstance(value, float | int):
            if negate:
                value = -value
            return value
        return 0.0


    # ---------------------------------------------------------------------------------------------
    def _Convert_Str_To_Float(self, Value):
        self.Dummy = 0
        if Value is None:
            return 0.0
        try:
            # Questo copre int, float e stringhe numeriche pulite (es: "123.45")
            return float(Value)
        except (ValueError, TypeError):
            # Se la conversione fallisce (es: c'è del testo o simboli strani)
            return 0.0
        finally:
            pass

# =================================================================================================